import os
import difflib
from Bio import pairwise2
from Bio.pairwise2 import format_alignment
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from Bio import SeqIO
from itertools import product
f= open('sequenceGenerated.doc','wt')

itext1 = open('SOD1_seq_ensemble.txt','r')
# itext2 = open('test_seq.txt', 'r')

# itext1 = str(itext1.read())
# itext1 = str(itext1[113:124])
# itext2 = str(itext2.read())
#
# alignments = pairwise2.align.globalms(itext1, itext2, 2, -1,-2,-.01)
# print(alignments)



wb = Workbook()
file = load_workbook('CanineHD_B_31_only.xlsx')
ws = file.active
Text1 = str(itext1.read())
Text1 = str(Text1[113:124])
results = []
tcheck = 0
for z in range(1,10): #2893

    cell_obj = ws.cell(row=int(z), column=17)
    Text2 = str(cell_obj.value)
    alignments = pairwise2.align.globalms(Text1, Text2, 1, -1, -1, -.01)
    f.write(format_alignment(*alignments[0]))



    timer = int(z)/2893
    timer = timer*100
    timer = round(timer,1)

    # if tcheck == timer:
    #     continue
    # else:
    #     if timer % 5 == 0:
    #
    #         print(str(timer)+'% done')
    #         tcheck = timer
    # if diff >= 5:
    #     results.append(str(diff)+'% match in row '+str(z))
    #     # print(str(diff)+'% match in row '+str(z))

f.close()